import cv2
from tkinter import *
import tkinter as tk
from PIL import Image, ImageTk
from pyzbar.pyzbar import decode
import datetime
import os
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook, load_workbook
from PIL import Image as PILImage


from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as PXIMAGE
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

strangedata =  ''
strangedata2 = ''



def get_user_info_by_id(user_id):
    try:
        connection = sqlite3.connect('user_data.db')
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
        user_info = cursor.fetchone()
        connection.close()
        return user_info

    except sqlite3.Error as e:
        print(f"Error accessing the database: {e}")
        return None
    
class QRCodeScanner:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Code Scanner")

        self.video_source = 0
        self.vid = cv2.VideoCapture(self.video_source)

        self.canvas = Canvas(root, width=self.vid.get(cv2.CAP_PROP_FRAME_WIDTH), height=self.vid.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self.canvas.pack()

        self.label = Label(root, text="QR Code will be displayed here", font=("Helvetica", 16))
        self.label.pack(pady=10)


        self.btn_start = Button(root, text="Start", command=self.start_scanning)
        self.btn_start.pack(side=LEFT, padx=10)

        self.btn_stop = Button(root, text="Stop", command=self.stop_scanning)
        self.btn_stop.pack(side=LEFT, padx=10)

        self.photo = None
        self.is_scanning = False
        self.update()

        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.today_date = datetime.datetime.now().strftime("%Y-%m-%d")
        self.excel_folder = f"excel_files/{self.today_date}"
        self.image_folder = f"captured_images/{self.today_date}"

        self.create_folders()

        self.saved_qr_codes = set()

    def start_scanning(self):
        self.is_scanning = True

    def stop_scanning(self):
        self.is_scanning = False

    def update(self):
        global strangedata
        if self.is_scanning:
            
            ret, frame = self.vid.read()

            if ret:
                qr_data = self.decode_qr_code(frame)
                strangedata = str(qr_data)
                if qr_data:
                    user_info = self.get_user_info_from_qr_data(qr_data)
                    if user_info:
                        self.label.config(text=f"User ID: {user_info[0]}, Name: {user_info[1]}")
                        result = get_user_info_by_id(user_info[0])
                        if result:
                            self.process_qr_data(qr_data, frame, user_info, f"{result[2]}", f"{result[3]}", f"{result[4]}")
                        else:
                            self.capture_and_save_image("unknown", frame)
                            print("User not found.")
                    else:
                        self.capture_and_save_image("unknown", frame)
                self.photo = self.convert_frame_to_photo(frame)
                self.canvas.create_image(0, 0, image=self.photo, anchor=NW)

        self.root.after(10, self.update)

    def decode_qr_code(self, frame):
        try:
            decoded_objects = decode(frame)
            if decoded_objects:
                return decoded_objects[0].data.decode("utf-8")
            else:
                return ""
        except Exception as e:
            print("Error decoding QR code:", e)
            return ""

    def get_user_info_from_qr_data(self, qr_data):
        try:
            user_id = None
            user_name = None

            key_value_pairs = qr_data.split(",")

            for pair in key_value_pairs:
                key, value = pair.split(":")
                key = key.strip().lower()

                if key == "id":
                    user_id = value.strip()
                elif key == "name":
                    user_name = value.strip()

            return (user_id, user_name) if user_id else None
        except Exception as e:
            print("Error extracting user information from QR data:", e)
            return None

    def process_qr_data(self, qr_data, frame, user_info, image_pat, email, stage):
        user_id, user_name = user_info

        if user_id not in self.saved_qr_codes:
            # Save user information to Excel
            self.save_to_excel(user_id, user_name, frame, image_pat, email, stage)
            self.saved_qr_codes.add(user_id)

        if not self.is_known_user(user_id):
            self.capture_and_save_image(user_id, frame)

    def is_known_user(self, user_id):
        try:
            connection = sqlite3.connect('user_data.db')
            cursor = connection.cursor()
            cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
            result = cursor.fetchone()
            return result is not None
        except sqlite3.Error as e:
            print("Error checking if user is known:", e)
            return False

    def save_to_excel(self, user_id, user_name, frame, image_pat, email, stage):
        excel_file = f"{self.excel_folder}/user_data.xlsx"
        image_path = f"storge/lastimage.png"
        cv2.imwrite(image_path, cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        timesign = datetime.datetime.now().strftime('%H:%M:%S')


        def insert_data_and_images(file_path, id, fullname, signin_time, original_image_path, selfie_image_path, email, stage):
            try:
             
                wb = load_workbook(file_path)
                sheet = wb.active
            except FileNotFoundError:
                wb = Workbook()
                sheet = wb.active
                for row in range(2, 100):
                    sheet.row_dimensions[row].height = 93
                    for col in range(1, 10):  
                        col_letter = get_column_letter(col)
                        sheet.column_dimensions[col_letter].width = 17
                sheet['A1'] = 'ID'
                sheet['B1'] = 'Fullname'
                sheet['C1'] = 'Signin Time'
                sheet['D1'] = 'Email'
                sheet['E1'] = 'Stage'
                sheet['F1'] = 'Original Image'
                sheet['G1'] = 'Selfie Image'

            row = sheet.max_row + 1

            sheet[f'A{row}'] = id
            sheet[f'B{row}'] = fullname
            sheet[f'C{row}'] = signin_time
            sheet[f'D{row}'] = email
            sheet[f'E{row}'] = stage


            for col in range(1, 10):
                col_letter = get_column_letter(col)
                sheet.column_dimensions[col_letter].width = 17

            original_img = PILImage.open(original_image_path)
            original_img.thumbnail((120, 120))  # Resize the image
            original_img.save('original_thumbnail.png')

            original_img_obj = PXIMAGE('original_thumbnail.png')
            original_img_obj.width = original_img.width
            original_img_obj.height = original_img.height
            original_img_obj.anchor = f'F{row}'  # Set the anchor to the cell

            sheet.add_image(original_img_obj)

            selfie_img = PILImage.open(selfie_image_path)
            selfie_img.thumbnail((120, 120))  # Resize the image
            selfie_img.save('selfie_thumbnail.png')

            selfie_img_obj = PXIMAGE('selfie_thumbnail.png')
            selfie_img_obj.width = selfie_img.width
            selfie_img_obj.height = selfie_img.height
            selfie_img_obj.anchor = f'G{row}' 

            sheet.add_image(selfie_img_obj)

            wb.save(file_path)
        insert_data_and_images(excel_file, user_id, user_name, timesign, image_pat, 'storge/lastimage.png', email, stage)

    def capture_and_save_image(self, user_id, frame):
        global strangedata2
        try:
            user_folder = f"{self.image_folder}"
            os.makedirs(user_folder, exist_ok=True)

            current_time = datetime.datetime.now().strftime('%H%M%S')
            
            if strangedata != strangedata2:
                self.label.config(text=f"XXXXX : Your unknown persone")
                image_path = f"{user_folder}/Unknown_{current_time}.png"
                cv2.imwrite(image_path, cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                strangedata2 = strangedata
                
        except Exception as e:
            print("Error capturing and saving image:", e)

    def convert_frame_to_photo(self, frame):
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        photo = ImageTk.PhotoImage(image=Image.fromarray(frame))

        return photo

    def on_close(self):
        self.vid.release()
        self.conn.close()
        self.root.destroy()

    def create_folders(self):
        os.makedirs(self.excel_folder, exist_ok=True)
        os.makedirs(self.image_folder, exist_ok=True)

if __name__ == "__main__":
    root = Tk()
    app = QRCodeScanner(root)
    root.mainloop()
